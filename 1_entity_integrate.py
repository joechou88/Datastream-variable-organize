import re
import os
import sys
import traceback
from datetime import datetime
import config
import integrator
from openpyxl import load_workbook

os.makedirs(config.ENTITY_OUTPUT_FOLDER, exist_ok=True)

log_file = open(config.ENTITY_LOG_FILE, "w", encoding="utf-8")
sys.stdout = integrator.Tee(sys.stdout, log_file)
sys.stderr = integrator.Tee(sys.stderr, log_file)

print("="*60)
print("Entity Integration Log")
print("Start Time:", datetime.now())
print("="*60)

def check_request_table(wb, fname, company_no, start, end):
    ws = wb[config.REQUEST_SHEET]
    expected_series = f"FDEALL{company_no}"

    start_year = int(start)
    end_year = int(end) if end else start_year
    expected_years = list(range(start_year, end_year + 1))

    row = 7
    year_idx = 0

    while ws[f"E{row}"].value not in (None, ""):
        # ===== E 欄：公司組數檢查 =====
        if ws[f"E{row}"].value != expected_series:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE E{row} = {ws[f'E{row}'].value}，"
                f"預期 {expected_series}"
            )
            return False

        # ===== G 欄：年份檢查 =====
        if year_idx >= len(expected_years):
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE 年份列數超出檔名範圍（從 G{row} 開始）"
            )
            return False

        raw_year = ws[f"G{row}"].value
        expected_year = expected_years[year_idx]

        try:
            cell_year = int(str(raw_year).strip())
        except Exception:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE G{row} = {raw_year}，"
                f"無法解析為年份"
            )
            return False

        if cell_year != expected_year:
            print(
                f"⚠️ 跳過: {fname} REQUEST_TABLE G{row} = {cell_year}，"
                f"預期 {expected_year}（與檔名年份不一致）"
            )
            return False

        row += 1
        year_idx += 1

    # ===== 列數反向檢查（避免少一年）=====
    if year_idx != len(expected_years):
        raise ValueError(
            f"{fname} REQUEST_TABLE 年份列數不足，"
            f"預期 {len(expected_years)} 列，實際 {year_idx} 列"
        )

def get_request_table_value(ws, col, start_row=7):
    """
    讀 REQUEST_TABLE 連續非空值（從 N7 / O7 開始）
    回傳 list[int]
    """
    values = []
    row = start_row
    while ws[f"{col}{row}"].value not in (None, ""):
        try:
            values.append(int(ws[f"{col}{row}"].value))
        except Exception:
            values.append(None)
        row += 1
    return values

def validate_wb(wb, fname, company_no, start, end, years):
    # ===== 確定 REQUEST_TABLE 存在 =====
    if config.REQUEST_SHEET not in wb.sheetnames:
        raise ValueError(f"{fname} 缺少 REQUEST_TABLE")
    
    # ===== 檢查 檔名和 REQUEST_TABLE 的 Series 一致 =====
    check_request_table(wb, fname, company_no, start, end)
    
    # ===== 檢查 檔名和 工作表數量 一致 =====
    data_sheets = [s for s in wb.sheetnames if s != config.REQUEST_SHEET]
    if len(data_sheets) < years:
        raise ValueError(
            f"{fname} 工作表數量不足，預期 {years} 張，實際 {len(data_sheets)}"
        )

def print_sheet_shapes(wb, fname, skip_sheet=config.REQUEST_SHEET):
    for ws_name in wb.sheetnames:
        if ws_name == skip_sheet:
            continue
        ws = wb[ws_name]
        rows = integrator.actual_rows(ws)
        cols = integrator.actual_cols(ws)
        print(f"{fname} 🔹 工作表: {ws_name}, "
              f"shape: {rows} rows x {cols} columns")

def append_sheet_rows(target_ws, source_ws, fname_only, base_cols_by_year, src_cols_by_year, year_idx):
    """
    將 source_ws 的資料接到 target_ws 後面
    - 只允許欄位數一致
    - 不一致時印出警告，但仍跳過 append
    """
    target_cols = integrator.actual_cols(target_ws)
    source_cols = integrator.actual_cols(source_ws)

    if target_cols != source_cols:
        print(
            f"❌ COLS 不一致 | "
            f"{os.path.basename(target_ws.parent.properties.title)} "
            f"O{7+year_idx}={base_cols_by_year[year_idx]} | "
            f"{fname_only} O{7+year_idx}={src_cols_by_year[year_idx]}"
        )
        return False  # 不 append

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        # 跳過空白列
        if not any(cell is not None for cell in row):
            continue
        target_ws.append(row)
    
    return True

try:
    try:
        expected_company_count = int(
            input("🧩 請輸入每個國家預期的公司群數（例如 8）: ").strip()
        )
        if expected_company_count < 1:
            raise ValueError
    except ValueError:
        print("❌ 請輸入大於等於 1 的整數")
        exit(1)

    files = [
        f for f in os.listdir(config.ENTITY_INPUT_FOLDER)
        if f.endswith((".xlsx", ".xlsm"))
    ]

    groups = {}
    key_to_outname = {}

    for f in files:
        info = integrator.parse_filename(f)
        if not info:
            continue
        key = (
            info["country"],
            info["start"],
            info["end"],
            info["suffix"]
        )
        groups.setdefault(key, []).append((int(info["company"]), f))

        if key not in key_to_outname:
            out_name = f"{info['country']}-{info['start']}{'-'+info['end'] if info['end'] else ''}{info['suffix']}.xlsx"
            key_to_outname[key] = out_name

    missing_company_report = []
    existing_outputs = []

    for (country, start, end, suffix) in groups.keys():
        out_name = key_to_outname[(country, start, end, suffix)]
        out_path = os.path.join(config.ENTITY_OUTPUT_FOLDER, out_name)

        if os.path.exists(out_path):
            existing_outputs.append(out_path)

    if existing_outputs:
        print("\n⚠️  以下輸出檔案已存在，將被覆蓋：")
        for p in existing_outputs:
            print(f"   - {p}")

        ans = input("\n是否同意刪除並全部重生？(y/N): ").strip().lower()

        if ans not in ("y", "yes"):
            print(
                "\n❌ 已取消執行。\n"
                "請自行到 ./data-split-by-variable 刪除上述檔案後再重新執行。"
            )
            exit(1)

        for p in existing_outputs:
            os.remove(p)
            print(f"🗑 已刪除：{p}")
        print(f"\n========================\n")

    for (country, start, end, suffix), items in groups.items():    
        companies = {company: fname for company, fname in items}
        actual_companies = set(companies.keys())
        expected_companies = set(range(1, expected_company_count + 1))
        missing_companies = sorted(expected_companies - actual_companies)

        if missing_companies:
            missing_company_report.append({
                "country": country,
                "period": f"{start}{'-' + end if end else ''}{suffix}",
                "missing": missing_companies
            })

        if 1 not in companies:
            raise ValueError(
                f"缺少 company=1，無法合併：{country}-{start}{'-'+end if end else ''}{suffix}"
            )

        base_company = 1
        base_file = os.path.join(config.ENTITY_INPUT_FOLDER, companies[1])

        wb_base = load_workbook(base_file, data_only=True)
        years = 1 if end is None else int(end) - int(start) + 1
        merged_rows_by_year = [0] * years

        validate_wb(wb_base, base_file, base_company, start, end, years)
        print_sheet_shapes(wb_base, companies[1])

        base_sheet_names = [s for s in wb_base.sheetnames if s != config.REQUEST_SHEET]

        for year_idx, ws_name in enumerate(base_sheet_names):
            ws = wb_base[ws_name]
            rows = integrator.actual_rows(ws)
            merged_rows_by_year[year_idx] += rows + 1

        for company in sorted(companies):
            if company == 1:
                continue
            fname_only = companies[company]
            fname = os.path.join(config.ENTITY_INPUT_FOLDER, fname_only)
            wb_src = load_workbook(fname, data_only=True)

            ws_req_base = wb_base[config.REQUEST_SHEET]
            ws_req_src = wb_src[config.REQUEST_SHEET]

            base_cols_by_year = get_request_table_value(ws_req_base, "O")
            src_cols_by_year = get_request_table_value(ws_req_src, "O")

            validate_wb(wb_src, fname, company, start, end, years)

            for ws_name in wb_base.sheetnames:
                if ws_name == config.REQUEST_SHEET:
                    continue

                ws_base = wb_base[ws_name]
                ws_src = wb_src[ws_name]

                rows = integrator.actual_rows(ws_src)
                cols = integrator.actual_cols(ws_src)

                year_idx = list(
                    s for s in wb_base.sheetnames if s != config.REQUEST_SHEET
                ).index(ws_name)

                print(
                    f"{fname_only} 🔹 工作表: {ws_name}, "
                    f"shape: {rows} rows x {cols} columns"
                )

                appended = append_sheet_rows(ws_base, ws_src, fname_only, base_cols_by_year, src_cols_by_year, year_idx)

                if appended:
                    merged_rows_by_year[year_idx] += rows

        out_name = key_to_outname[(country, start, end, suffix)]
        out_path = os.path.join(config.ENTITY_OUTPUT_FOLDER, out_name)

        print(f"\n📊 {out_name} 最終合併後 sheet shape：")
        print_sheet_shapes(wb_base, out_name)
        
        ws_req = wb_base[config.REQUEST_SHEET]
        data_sheets = [s for s in wb_base.sheetnames if s != config.REQUEST_SHEET]

        for i, ws_name in enumerate(data_sheets):
            ws = wb_base[ws_name]
            rows = integrator.actual_rows(ws) + 1
            cols = integrator.actual_cols(ws)
            total = rows * cols

            ws_req[f"N{7+i}"].value = rows
            ws_req[f"O{7+i}"].value = cols
            ws_req[f"P{7+i}"].value = total

            print(
                f"🧮 {out_name} REQUEST_TABLE row {7+i}: "
                f"N={rows}, O={cols}, P={total}"
            )

        wb_base.save(out_path)
        print(f"✔ 輸出完成：{out_path}")
        print(f"\n========================\n")

    if missing_company_report:
        print("\n⚠️ 公司群數量警示（不影響輸出）")
        print("====================================")
        for item in missing_company_report:
            print(
                f"{item['country']}-{item['period']} "
                f"缺少公司群：{', '.join(map(str, item['missing']))}"
            )
    else:
        print("\n✅ 所有國家公司群數量皆符合預期")

except SystemExit:
    pass
except Exception as e:
    print("\n❌ 系統發生未預期錯誤：")
    traceback.print_exc()
finally:
    print("\nEnd Time:", datetime.now())
    print("="*60)
    log_file.close()

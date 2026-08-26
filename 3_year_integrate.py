import os
import pandas as pd
import re
import sys
import traceback
import config
import integrator
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook, Workbook

os.makedirs(config.YEAR_OUTPUT_FOLDER, exist_ok=True)

def extract_sheet_name(ref):
    return ref.split("!")[0].replace("'", "")

log_file = open(config.YEAR_LOG_FILE, "w", encoding="utf-8")

sys.stdout = integrator.Tee(sys.stdout, log_file)
sys.stderr = integrator.Tee(sys.stderr, log_file)

print("="*60)
print("Year Integration Log")
print("Start Time:", datetime.now())
print("="*60)

try:
    country_files = defaultdict(list)

    for f in os.listdir(config.VARIABLE_OUTPUT_FOLDER):
        if f.lower().endswith((".xlsm", ".xlsx")):
            filename = integrator.parse_filename(f)
            if filename:
                country_files[filename["country"]].append(f)

    code_df = pd.read_excel(config.COUNTRY_CODE_INPUT)
    code_df["Country_name"] = code_df["Country_name"].str.strip()

    country_code_map = code_df.set_index("Country_name").to_dict(orient="index")

    for country, files in country_files.items():
        print(f"\nProcessing {country}...")

        country = country.replace("-", " ")

        records = []
        header = None
        expected_cols = None
        year_col_count = {}

        for fname in files:
            path = os.path.join(config.VARIABLE_OUTPUT_FOLDER, fname)
            wb = load_workbook(path, data_only=True)

            if "REQUEST_TABLE" not in wb.sheetnames:
                wb.close()
                print(f"❌ 缺少 REQUEST_TABLE！略過國家：{country}")
                continue

            req_ws = wb["REQUEST_TABLE"]
            row = 7

            # REQUEST_TABLE 從第 7 列一直往下讀，讀到空白就停
            while True:
                filename = integrator.parse_filename(f)
                if filename:
                    start_year = filename["start"]
                    end_year = filename["end"] if filename["end"] else start_year
                    file_years = set(range(start_year, end_year + 1))
                else:
                    file_years = set()

                year = req_ws[f"G{row}"].value
                ref = req_ws[f"K{row}"].value

                if year is None or ref is None:
                    break

                year = int(year)

                # ====== 檔名 vs REQUEST_TABLE 年份檢查 ======
                if year not in file_years:
                    print(
                        f"❌ 年份不一致｜檔名: {fname} "
                        f"| REQUEST_TABLE 年份: {year} "
                        f"| 檔名年份: {sorted(file_years)} → 已跳過"
                    )
                    row += 1
                    continue

                if config.START_YEAR <= year <= config.END_YEAR:
                    sheet_name = extract_sheet_name(ref)

                    if sheet_name in wb.sheetnames:
                        src_ws = wb[sheet_name]
                        raw_rows = list(src_ws.iter_rows(values_only=True))
                        rows = [
                            r for r in raw_rows
                            if any(cell is not None for cell in r) # Excel 被更動過會殘留「看不見的空白列」，需自動丟棄
                        ]

                        # 若有丟棄空白列，印警告
                        if len(rows) < len(raw_rows):
                            print(
                                f"⚠ 警告｜{country} {year} 年："
                                f"工作表包含 {len(raw_rows)-len(rows)} 列殘留空白列，已自動移除"
                            )

                        if not rows:
                            row += 1
                            continue

                        # ====== 確保同一國家變數數量都一樣 ======
                        # 優先檢查 REQUEST_TABLE O 欄
                        # 備援：實際去數後面工作表欄位 - 1
                        cols_value = req_ws[f"O{row}"].value
                        rows_value = req_ws[f"N{row}"].value
                        print(
                            f"國家: {country} | 年份: {year} "
                            f"| O欄(cols_value) = {cols_value} "
                            f"| N欄(rows_value) = {rows_value}"
                            f"| 工作表列數={len(rows)}"
                        )

                        if isinstance(cols_value, int):
                            n_cols = cols_value
                        else:
                            n_cols = src_ws.max_column
                        number_of_variables = n_cols - 1 # 排除 Type (DSCD)

                        year_col_count[year] = n_cols

                        if expected_cols is None:
                            expected_cols = number_of_variables # 紀錄該國第一年變數數量
                        else:
                            if number_of_variables != expected_cols:
                                print(f"❌ 變數數量不一致，已略過該年份！國家：{country} 年份：{year}")
                                print(f"  期望變數數量（不含 Type/DSCD）：{expected_cols}")
                                print(f"  年份 {year} 變數數量：{number_of_variables}")
                                print("  各年變數數量（不含 Type/DSCD）：")
                                for y, c in year_col_count.items():
                                    print(f"   - {y}: {c-1}")
                                row += 1
                                continue # 跳回 while True 的開頭，跑下一年
                        # ====== 檢查結束 ======

                        if header is None:
                            header = ["YEAR"] + list(rows[0]) # 第一次跑該國時，紀錄欄位名稱

                        for data_row in rows[1:]:
                            records.append([year] + list(data_row)) # 把「某一年的資料」一筆一筆加進 MASTER_TABLE

                row += 1

            wb.close()

        if len(records) == 0:
            print(f"  ⚠ {country} 無有效資料，略過")
            continue

        # Add country info
        code_info = country_code_map.get(country, {"Country_code": "", "Country_code2": ""})
        country_code = code_info.get("Country_code", "")
        country_code2 = code_info.get("Country_code2", "")

        # 調整 header，把三欄插到 YEAR 之後
        new_header = header[:1] + ["COUNTRY", "COUNTRY_CODE", "COUNTRY_CODE2"] + header[1:]

        # 調整每筆資料
        new_records = []
        for row in records:
            new_row = row[:1] + [country, country_code, country_code2] + row[1:]
            new_records.append(new_row)

        # ---- 依年份升冪排序 ----
        new_records.sort(key=lambda x: x[0])

        # ---- 取得實際年份範圍 ----
        years_present = sorted({row[0] for row in new_records})

        # ---- 檢查年份完整性 ----
        required_years = set(range(config.START_YEAR, config.END_YEAR + 1))
        missing_required = sorted(required_years - set(years_present))

        # ---- 檢查年份連續性 ----
        min_year = min(years_present)
        max_year = max(years_present)
        full_range = set(range(min_year, max_year + 1))
        missing_continuous = sorted(full_range - set(years_present))

        # ---- 若任一條件不符合就跳過 ----
        if missing_required or missing_continuous:
            msg = f"⚠ {country} 資料不完整或不連續｜實際年份: {years_present}"
            if missing_required:
                msg += f" | 年份不完整，缺少: {', '.join(str(y) for y in missing_required)}"
            if missing_continuous:
                msg += f" | 年份不連續，缺少: {', '.join(str(y) for y in missing_continuous)}"
            print(msg)
            continue  # 跳過該國家，不輸出

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "MASTER_TABLE"

        out_ws.append(new_header)
        for r in new_records:
            out_ws.append(r)

        out_path = os.path.join(
            config.YEAR_OUTPUT_FOLDER, f"{country}-{min_year}-{max_year}.xlsx"
        )
        out_wb.save(out_path)

        print(f"  ✔ 輸出完成: {out_path}，共 {len(new_records)} 筆資料")

    print("=== 全部國家彙整完成 ===")

except SystemExit:
    pass
except Exception as e:
    print("\n❌ 系統發生未預期錯誤：")
    traceback.print_exc()
finally:
    print("\nEnd Time:", datetime.now())
    print("="*60)
    log_file.close()

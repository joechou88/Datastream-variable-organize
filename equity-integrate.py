import re
import os
from openpyxl import load_workbook

# ================== 設定 ==================
INPUT_FOLDER = "data-split-by-equity"
OUTPUT_FOLDER = "data-split-by-variable"
REQUEST_SHEET = "REQUEST_TABLE"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ================== 檔名解析 ==================
pattern = re.compile(
    r"""
    (?P<country>[A-Za-z]+)
    (?P<company>\d+)
    -
    (?P<start>\d{4})
    (?:-(?P<end>\d{4}))?
    (?P<suffix>[A-Za-z]+)
    """,
    re.VERBOSE
)

def parse_filename(fname):
    name = os.path.splitext(fname)[0]
    m = pattern.fullmatch(name)
    return m.groupdict() if m else None

def check_request_table(wb, company_no, years):
    ws = wb[REQUEST_SHEET]
    expected = f"FDEALL{company_no}"

    if years == 1:
        return ws["E7"].value == expected
    else:
        for r in range(7, 7 + years):
            if ws[f"E{r}"].value != expected:
                return False
        return True

def validate_wb(wb, fname, company_no, years):
    # ===== 確定 REQUEST_TABLE 存在 =====
    if REQUEST_SHEET not in wb.sheetnames:
        raise ValueError(f"{fname} 缺少 REQUEST_TABLE")
    
    # ===== 檢查 檔名和 REQUEST_TABLE 的 Series 一致 =====
    if not check_request_table(wb, company_no, years):
        raise ValueError(f"REQUEST_TABLE 的 Series 與檔名不符：{fname}")
    
    # ===== 檢查 檔名和 工作表數量 一致 =====
    if len(wb.sheetnames) < years:
        raise ValueError(
            f"{fname} 工作表數量不足，預期 {years} 張，實際 {len(wb.sheetnames)}"
        )

def sheet_shape(ws):
    rows = ws.max_row - 1  # 扣掉 header
    cols = ws.max_column
    return rows, cols

# ================== row append ==================
def append_sheet_rows(target_ws, source_ws):
    start_row = target_ws.max_row + 1
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        target_ws.append(row)

# ================== 主流程 ==================
try:
    expected_company_count = int(
        input("🧩 請輸入每個國家預期的公司群數（例如 8）: ").strip()
    )
    if expected_company_count < 1:
        raise ValueError
except ValueError:
    print("❌ 請輸入大於等於 1 的整數")
    exit(1)

files = [
    f for f in os.listdir(INPUT_FOLDER)
    if f.endswith((".xlsx", ".xlsm"))
]

groups = {}
key_to_outname = {}

for f in files:
    info = parse_filename(f)
    if not info:
        continue
    key = (
        info["country"],
        info["start"],
        info["end"],
        info["suffix"]
    )
    groups.setdefault(key, []).append((int(info["company"]), f))

    if key not in key_to_outname:
        out_name = f"{info['country']}-{info['start']}{'-'+info['end'] if info['end'] else ''}{info['suffix']}.xlsx"
        key_to_outname[key] = out_name

missing_company_report = []
existing_outputs = []

for (country, start, end, suffix) in groups.keys():
    out_name = key_to_outname[(country, start, end, suffix)]
    out_path = os.path.join(OUTPUT_FOLDER, out_name)

    if os.path.exists(out_path):
        existing_outputs.append(out_path)

if existing_outputs:
    print("\n⚠️  以下輸出檔案已存在，將被覆蓋：")
    for p in existing_outputs:
        print(f"   - {p}")

    ans = input("\n是否同意刪除並全部重生？(y/N): ").strip().lower()

    if ans not in ("y", "yes"):
        print(
            "\n❌ 已取消執行。\n"
            "請自行到 ./data-split-by-variable 刪除上述檔案後再重新執行。"
        )
        exit(1)

    for p in existing_outputs:
        os.remove(p)
        print(f"🗑 已刪除：{p}")

for (country, start, end, suffix), items in groups.items():    
    companies = {company: fname for company, fname in items}
    actual_companies = set(companies.keys())
    expected_companies = set(range(1, expected_company_count + 1))
    missing_companies = sorted(expected_companies - actual_companies)

    if missing_companies:
        missing_company_report.append({
            "country": country,
            "period": f"{start}{'-' + end if end else ''}{suffix}",
            "missing": missing_companies
        })

    # ===== 嚴格檢查：一定要有 company = 1 作為模板 =====
    if 1 not in companies:
        raise ValueError(
            f"缺少 company=1，無法合併：{country}-{start}{'-'+end if end else ''}{suffix}"
        )

    base_company = 1
    base_file = os.path.join(INPUT_FOLDER, companies[1])

    wb_base = load_workbook(base_file, keep_vba=False)
    years = 1 if end is None else int(end) - int(start) + 1

    validate_wb(wb_base, base_file, base_company, years)

    for company in sorted(companies):
        if company == 1:
            continue
        fname_only = companies[company]
        fname = os.path.join(INPUT_FOLDER, fname_only)
        wb_src = load_workbook(fname, data_only=True)

        validate_wb(wb_src, fname, company, years)

        for i in range(years):
            ws_name = wb_base.sheetnames[i]
            ws_src = wb_src[ws_name]

            rows, cols = sheet_shape(ws_src)

            print(
                f"{fname_only} 🔹 工作表: {ws_name}, "
                f"shape: {rows} rows x {cols} columns"
            )

            append_sheet_rows(
                wb_base[ws_name],
                ws_src
            )

    out_name = key_to_outname[(country, start, end, suffix)]
    out_path = os.path.join(OUTPUT_FOLDER, out_name)

    wb_base.save(out_path)
    print(f"✔ 輸出完成：{out_path}")

if missing_company_report:
    print("\n⚠️ 公司群數量警示（不影響輸出）")
    print("====================================")
    for item in missing_company_report:
        print(
            f"{item['country']}-{item['period']} "
            f"缺少公司群：{', '.join(map(str, item['missing']))}"
        )
else:
    print("\n✅ 所有國家公司群數量皆符合預期")

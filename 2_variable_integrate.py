import os
import re
import sys
import traceback
import config
import integrator
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict

os.makedirs(config.VARIABLE_OUTPUT_FOLDER, exist_ok=True)

def get_expected_output_files(parsed, country_year_spans):
    outputs = {}  # out_path -> (country, year_label)

    for country, spans in country_year_spans.items():
        is_consistent, year_span_list = check_year_span_consistency(
            country, spans
        )
        if not is_consistent:
            continue

        for start_year, end_year in year_span_list:
            year_label = (
                f"{start_year}"
                if start_year == end_year
                else f"{start_year}-{end_year}"
            )
            fname = f"{country}-{year_label}.xlsx"
            out_path = os.path.join(config.VARIABLE_OUTPUT_FOLDER, fname)
            outputs[out_path] = (country, year_label)

    return outputs

def create_output_file(country, start_year, end_year):
    year_label = (
        f"{start_year}"
        if start_year == end_year
        else f"{start_year}-{end_year}"
    )

    fname = f"{country}-{year_label}.xlsx"

    out_path = os.path.join(config.VARIABLE_OUTPUT_FOLDER, fname)
    
    files = [f for f in os.listdir(config.ENTITY_OUTPUT_FOLDER) if f.endswith((".xlsx", ".xlsm"))]

    try:
        template_fname = find_excel_file(country, start_year, "A", files)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"❌ 找不到 {country}-{start_year}A.xlsx 或 .xlsm 作為模板"
        )

    template_path = os.path.join(config.ENTITY_OUTPUT_FOLDER, template_fname)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"找不到檔案：{template_path}")
    
    wb = load_workbook(template_path)

    if "REQUEST_TABLE" not in wb.sheetnames:
        raise ValueError(f"{template_fname} 中沒有 REQUEST_TABLE 工作表")
    
    wb.save(out_path)
    
    return out_path

def find_excel_file(country, start_year, var_tag, files):
    """
    找出指定國家、年份、變數的檔案（A/B/C...）
    支援單年或跨年
    """
    # 精確匹配 country-startyear(-endyear)var_tag
    pattern = re.compile(
        rf"^{re.escape(country)}-{start_year}(?:-\d{{4}})?[A-Z]*{var_tag}[A-Z]*\.(xlsx|xlsm)$"
    )
    candidates = [f for f in files if pattern.match(f)]

    if not candidates:
        raise FileNotFoundError(
            f"❌ 找不到 {country}-{start_year}{var_tag}.xlsx 或 .xlsm"
        )

    # 如果剛好有兩個（理論上不應該），優先用 .xlsx
    candidates.sort(key=lambda x: x.endswith(".xlsm"))
    return candidates[0]


def check_year_span_consistency(country, year_spans):
    """
    1) 將所有年份標準化，單一年 -> (year, year)
    2) 依開始年排序，找出各個 year_span
    3) 確保同一 year_span 的 A/B/C/... 年段完全一致
    4) 後續 year_span 不能重疊先前 year_span 年份
    回傳：
        - is_consistent: True/False
        - year_span_list: list of (start_year, end_year)
    """
    # 標準化：只有一年 -> (year, year)
    normalized_year_span = [(s, s) if e is None else (s, e) for s, e in year_spans]

    # 依開始年排序
    normalized_year_span = sorted(normalized_year_span, key=lambda x: x[0])

    year_span_list = []
    current_start, current_end = normalized_year_span[0]

    for s, e in normalized_year_span[1:]:
        if s <= current_end:  # 屬於同一個 year_span
            if (s, e) != (current_start, current_end):
                print(f"\n🚨 {country}：同一 year_span A/B/C 年段不一致")
                print(f"  Expected：{current_start}-{current_end}")
                print(f"  Found：{s}-{e}")
                return False, None
        else:   # 新 year_span
            year_span_list.append((current_start, current_end))
            current_start, current_end = s, e

    year_span_list.append((current_start, current_end))  # 加最後一個 year_span

    # 檢查 year_span 之間不重疊
    for i in range(1, len(year_span_list)):
        prev_s, prev_e = year_span_list[i-1]
        curr_s, curr_e = year_span_list[i]
        if curr_s <= prev_e:
            print(f"\n🚨 {country}：與前一個 year_span 重疊")
            print(f"  前一個 year_span：{prev_s}-{prev_e}")
            print(f"  當前 year_span：{curr_s}-{curr_e}")
            return False, None

    return True, year_span_list

def read_request_table(xls_path):
    """讀取 REQUEST_TABLE，回傳 dataframe（row=7 開始）"""
    return pd.read_excel(
        xls_path, sheet_name="REQUEST_TABLE", engine="openpyxl", header=None
    )

def get_sheet_for_year(req_df, year):
    """根據 REQUEST_TABLE 找到對應年份的工作表位置"""
    
    # 從 row7 開始抓 G欄（index=6）
    df_years = pd.to_numeric(req_df.iloc[6:, 6], errors='coerce')
    matches = df_years[df_years == year]

    if matches.empty:
        print("🔍 DEBUG：REQUEST_TABLE G 欄 'Start Date'（前 5 筆）內容如下：")
        print(df_years.head(5).tolist())
        raise ValueError(f"⚠️ REQUEST_TABLE 找不到年份 {year}")

    # 取第一個符合年份的列索引
    row_idx = matches.index[0]
    row_series = req_df.iloc[row_idx]

    sheet_ref = row_series[10]      # K欄
    expected_rows = row_series[13]  # N欄
    expected_cols = row_series[14]  # O欄

    # sheet_ref 形如: 工作表1'!$A$1
    sheet_name = sheet_ref.split("!") [0].replace("'", "")

    return sheet_name, int(expected_rows), int(expected_cols), row_idx + 1

def read_variable_data(xls_path, sheet_name):
    df = pd.read_excel(xls_path, sheet_name=sheet_name, engine="openpyxl")
    return df

def append_column(wb_out, df, sheet_name, variable_suffix):
    """
    Merge using Type in Column A as the primary key
    """

    # 讀取現有 sheet
    if sheet_name in wb_out.sheetnames:
        ws = wb_out[sheet_name]

        data = ws.values
        columns = next(data)
        base_df = pd.DataFrame(data, columns=columns)

    else:
        # 如果不存在，直接寫入
        base_df = pd.DataFrame()

    # 整理新資料
    df = df.copy()
    df.columns = df.columns.astype(str)

    if "Type" not in df.columns:
        raise ValueError("❌ 新資料沒有 Type 欄")
    
    df = df.dropna(subset=["Type"])     # 丟掉空值
    df["Type"] = df["Type"].astype(str).str.strip()     # 轉字串
    df = df[df["Type"] != ""]       # 丟掉空字串

    # 如果 base 是空
    if base_df.empty:
        merged_df = df

    else:
        if "Type" not in base_df.columns:
            raise ValueError("❌ 既有資料沒有 Type 欄")
        
        base_df = base_df.dropna(subset=["Type"])     # 丟掉空值
        base_df["Type"] = base_df["Type"].astype(str).str.strip()     # 轉字串
        base_df = base_df[base_df["Type"] != ""]       # 丟掉空字串
        
        # 清掉殘留欄位
        if "_order" in base_df.columns:
            base_df = base_df.drop(columns="_order")

        # 保留原順序
        base_df["_order"] = range(len(base_df))

        # 拿新資料「除了 Type 以外」的欄
        new_cols = [c for c in df.columns if c != "Type"]

        # ========= 公司差異分析 =========
        base_types = list(base_df["Type"])
        new_types = list(df["Type"])

        base_index_map = {t: i+2 for i, t in enumerate(base_types)}  # +2 因為 Excel 有表頭
        new_index_map = {t: i+2 for i, t in enumerate(new_types)}

        set_base = set(base_types)
        set_new = set(new_types)

        only_in_new = sorted(set_new - set_base)
        only_in_base = sorted(set_base - set_new)

        merged_df = pd.merge(
            base_df,
            df[["Type"] + new_cols],
            on="Type",
            how="outer",
            sort=False
        )

        # 排序：A 原順序在前，新公司排後
        merged_df = merged_df.sort_values("_order", na_position="last")
        merged_df = merged_df.drop(columns=["_order"])

        # 重新建立 index map
        final_index_map = {
            t: i+2 for i, t in enumerate(merged_df["Type"])
        }

        # -------- 新公司 --------
        for idx, company in enumerate(only_in_new):
            new_row_position = len(base_types) + idx + 2
            print(
                f"新公司 {company} 出現在 {sheet_name}{variable_suffix} 的第 {new_index_map[company]} 列，"
                f"加進 {sheet_name}A 的第 {final_index_map[company]} 列"
            )

        # -------- 少公司 --------
        for company in only_in_base:
            print(
                f"公司 {company} 出現在 {sheet_name}A 的第 {base_index_map[company]} 列，"
                f"但沒有出現在 {sheet_name}{variable_suffix}，"
                f"該公司 {variable_suffix} 組變數的值全部補 ."
            )

    # 將 NaN 轉為 "."
    merged_df = merged_df.fillna(".")

    # 清空舊 sheet
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])

    ws = wb_out.create_sheet(title=sheet_name)

    # 寫回
    for r in dataframe_to_rows(merged_df, index=False, header=True):
        ws.append(r)

def update_request_table(wb_out, src_path, out_path, excel_row, sheet_name):
    """
    先檢查 N 欄是否與來源檔一致，
    再以合併後的 sheet 實際資料計算 N/O/P，更新 REQUEST_TABLE，
    並印出加總過程
    """
    wb_src = load_workbook(src_path)

    ws_out = wb_out["REQUEST_TABLE"]
    ws_src = wb_src["REQUEST_TABLE"]

    # ========= 先檢查 N 欄 (Rows) =========
    N_COL = 14  # column N

    n_out = ws_out.cell(row=excel_row, column=N_COL).value
    n_src = ws_src.cell(row=excel_row, column=N_COL).value

    if n_out != n_src:
        print(
            f"⚠️ ROWS 不一致 | "
            f"{os.path.basename(out_path)} N{excel_row}={n_out} | "
            f"{os.path.basename(src_path)} N{excel_row}={n_src}"
        )

    # ========= 以合併後 sheet 的實際 shape 更新 =========
    ws_data = wb_out[sheet_name]

    rows = integrator.actual_rows(ws_data) + 1  # Count header as a column
    cols = integrator.actual_cols(ws_data)
    total = rows * cols

    ws_out[f"N{excel_row}"].value = rows
    ws_out[f"O{excel_row}"].value = cols
    ws_out[f"P{excel_row}"].value = total

    print(f"🧮 更新 REQUEST_TABLE {sheet_name} row {excel_row}: "
          f"N={rows}, O={cols}, P={total}")

log_file = open(config.VARIABLE_LOG_FILE, "w", encoding="utf-8")

sys.stdout = integrator.Tee(sys.stdout, log_file)
sys.stderr = integrator.Tee(sys.stderr, log_file)

print("="*60)
print("Variable Integration Log")
print("Start Time:", datetime.now())
print("="*60)

try:
    files = [f for f in os.listdir(config.ENTITY_OUTPUT_FOLDER) if f.endswith((".xlsx", ".xlsm"))]

    parsed = []
    for f in files:
        parsed.extend(integrator.parse_filename(f))

    # 依國家 -> 年度 -> 變數排序（A, B, C...）
    grouped = defaultdict(lambda: defaultdict(list))  # country -> year -> list of (var, fname)
    country_year_spans = defaultdict(list)
    
    for country, y1, y2, var, fname in parsed:
        country_year_spans[country].append((y1, y2))
        for y in range(y1, y2 + 1):
            grouped[country][y].append((var, fname))
    
    # 檢查之前是否已輸出過
    expected_outputs = get_expected_output_files(parsed, country_year_spans)

    existing_outputs = {
        path: meta
        for path, meta in expected_outputs.items()
        if os.path.exists(path)
    }

    if existing_outputs:
        print("\n⚠️ 發現以下輸出檔已存在 ./data：")
        for i, (path, (country, year_label)) in enumerate(existing_outputs.items(), 1):
            print(f"{i}. {country} ({year_label}) → {os.path.basename(path)}")

        while True:
            ans = input(
                "\n👉 是否【全部刪除】後重新產生？ (y/n): "
            ).strip().lower()

            if ans == "y":
                for path in existing_outputs:
                    print(f"🗑️ 刪除 {os.path.basename(path)}")
                    os.remove(path)
                break

            elif ans == "n":
                print(
                    "\n⏭️ 未刪除任何檔案。\n"
                    "請自行至 ./data 刪除欲重新產生的檔案後再執行。"
                )
                sys.exit(0)

            else:
                print("請輸入 y 或 n")

    for country, spans in grouped.items():

        # 先檢查該國所有檔案的年段是否一致
        is_consistent, year_span_list = check_year_span_consistency(
            country, country_year_spans[country]
        )
        if not is_consistent:
            continue   # 整個國家直接跳過，不輸出

        print(f"\n========== ▶ Processing {country} ==========")

        for start_year, end_year in year_span_list:
            print("\n" + "-" * 40)
            out_xlsx = create_output_file(country, start_year, end_year)
            if out_xlsx is None:
                continue   # 這個年度已做過，直接跳過
            wb_out = load_workbook(out_xlsx)
            skip_country = False

            # 篩選這個 block 的檔案
            block_files = [
                (y1, y2, var, fname)
                for parsed_country, y1, y2, var, fname in parsed
                if parsed_country == country and y1 >= start_year and y2 <= end_year
            ]
            block_files = sorted(block_files, key=lambda x: x[2])  # A/B/C 排序，以第一個最小字母先處理
            
            processed_files = set() # 記錄已處理 Excel

            for s, e, var, _ in block_files:
                fname = find_excel_file(country, s, var, files)
                if fname in processed_files:
                    continue  # 否則 Hong-Kong-2015CD 會被併 2 次
                processed_files.add(fname)  # 標記 Hong-Kong-2015CD 已處理

                src_path = os.path.join(config.ENTITY_OUTPUT_FOLDER, fname)
                vars_in_file = [v for _, _, v, f in block_files if f == fname]
                is_first_variable = ("A" in vars_in_file)
                print(f"📂 處理 {src_path}")

                req_df = read_request_table(src_path)

                for year in range(s, e+1):
                    try:
                        sheet_name, exp_rows, exp_cols, excel_row = get_sheet_for_year(req_df, year)
                        df = read_variable_data(src_path, sheet_name)
                        df_rows, df_cols = df.shape  # DataFrame 不含 header，會少一 row

                        actual_rows = df_rows + 1
                        actual_cols = df_cols

                        # 檢查尺寸
                        if actual_rows != exp_rows or actual_cols != exp_cols:
                            print(f"⚠️ {country}-{start_year}-{end_year}{var} rows/cols 不符"
                                f"   Expected: {exp_rows} rows x {exp_cols} cols\n"
                                f"   Actual:   {actual_rows} rows x {actual_cols} cols"
                            )
                            skip_country = True
                            break
                        else:
                            print(f"🔹 工作表: {sheet_name}, shape: {exp_rows} rows x {exp_cols} columns")

                        if is_first_variable:   # A 組變數作為模板，已經在新檔裡，skip
                            continue

                        append_column(
                            wb_out=wb_out,
                            df=df,
                            sheet_name=sheet_name,
                            variable_suffix=var
                        )

                        if not is_first_variable:
                            update_request_table(
                                wb_out=wb_out,
                                src_path=src_path,
                                out_path=out_xlsx,
                                excel_row=excel_row,
                                sheet_name=sheet_name
                            )
                    except Exception as e:
                        print(f"⚠️ ERROR: {e}")
                        skip_country = True
                        break   # 跳出 var 迴圈，外層會處理刪檔 + 換國
            
            wb_out.save(out_xlsx)   

            if skip_country:
                if os.path.exists(out_xlsx):
                    print(f"🗑️ 刪除檔案 {out_xlsx}")
                    os.remove(out_xlsx)
                break   # 跳出 year 迴圈 (略過後續年度)，換下一國

    print("🎉 所有國家/年度整合完成！")

except SystemExit:
    pass
except Exception as e:
    print("\n❌ 系統發生未預期錯誤：")
    traceback.print_exc()
finally:
    print("\nEnd Time:", datetime.now())
    print("="*60)
    log_file.close()
